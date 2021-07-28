import streamlit as st
import numpy as np
import pandas as pd
import openpyxl as pxl
pop = pd.read_excel('pop.xlsx')
exp = pxl.load_workbook('pop.xlsx')
sheet = exp['PCA']
stateList = []
st.title('Population Analysis (States - India)')
for i in range(2, sheet.max_row+1):
    if(sheet.cell(row=i, column=8).value != 'India'):
        stateList.append(sheet.cell(row=i, column=8).value)

# st.write(stateList)
updStateList = []
updStateList = list(set(stateList))
# st.write(updStateList)

# selected = st.selectbox('Select State', updStateList)
first_number = st.number_input("Enter the First Number")
second_number = st.number_input("Enter the Second Number")
# for i in range(2, sheet.max_row):
#     if((sheet.cell(row=i, column=8).value.lower() == selected.lower()) and (sheet.cell(row=i, column=9).value.lower() == 'total') and (sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value) <= second_number and (sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value) >= first_number):
#         st.write(f"State = {sheet.cell(row=i, column=8).value.title()}\n Total Male: {sheet.cell(row=i, column=12).value}\n Total Female: {sheet.cell(row=i, column=13).value}\n MF Ratio: {(sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value)}")
counter = 0
for i in range(2, sheet.max_row):
    if((sheet.cell(row=i, column=9).value.lower() == 'total') and (sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value) <= second_number and (sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value) >= first_number and (sheet.cell(row=i, column=8).value != 'India')):
        st.write(f"State = {sheet.cell(row=i, column=8).value.title()}\n Total Male: {sheet.cell(row=i, column=12).value}\n Total Female: {sheet.cell(row=i, column=13).value}\n MF Ratio: {(sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value)}")
        counter += 1
st.write(counter)

# for i in range(2, sheet.max_row):
#     if(((sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value) > 1.1) and (sheet.cell(row=i, column=9).value) == 'Total'):
#         st.write(f"State = {sheet.cell(row=i, column=8).value.title()} Total Male: {sheet.cell(row=i, column=12).value}, Total Female: {sheet.cell(row=i, column=13).value} MF Ratio: {(sheet.cell(row=i, column=12).value) / (sheet.cell(row=i, column=13).value)}")
